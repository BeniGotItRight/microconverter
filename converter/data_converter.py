"""Data conversions: CSV, Excel, PDF to Excel, Excel to PDF, JSON, XML, YAML."""

from pathlib import Path

import pandas as pd

# Try chardet for encoding detection
try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False


def _detect_encoding(file_path: Path) -> str:
    """Detect file encoding for CSV. Fallback to utf-8."""
    if not HAS_CHARDET:
        return "utf-8"
    with open(file_path, "rb") as f:
        raw = f.read(100000)
    result = chardet.detect(raw)
    return result.get("encoding") or "utf-8"


def _deduplicate_columns(columns: list) -> list:
    """Ensure column names are unique to avoid Pandas reindexing errors."""
    new_cols = []
    for i, col in enumerate(columns):
        name = str(col).strip() if col else f"Column_{i}"
        if name in new_cols:
            count = 1
            while f"{name}_{count}" in new_cols:
                count += 1
            name = f"{name}_{count}"
        new_cols.append(name)
    return new_cols


def csv_to_excel(
    input_path: Path,
    output_path: Path,
    encoding: str | None = None,
    delimiter: str | None = None,
) -> None:
    """
    Convert CSV to Excel. Handles encoding and delimiter for M-Pesa statements.
    """
    enc = encoding or _detect_encoding(input_path)
    try:
        df = pd.read_csv(input_path, encoding=enc, sep=delimiter or ",")
    except (UnicodeDecodeError, pd.errors.ParserError):
        # Fallback: try common encodings
        for fallback in ["utf-8", "latin-1", "iso-8859-1", "cp1252"]:
            try:
                df = pd.read_csv(input_path, encoding=fallback, sep=delimiter or ",")
                break
            except (UnicodeDecodeError, pd.errors.ParserError):
                continue
        else:
            raise ValueError("Could not decode CSV. Try specifying encoding in Advanced options.")

    df.to_excel(output_path, index=False, engine="openpyxl")
    _autofit_excel_columns(output_path)


def csv_to_pdf(input_path: Path, output_path: Path, encoding: str | None = None, delimiter: str | None = None) -> None:
    """Convert CSV to PDF as a formatted table."""
    enc = encoding or _detect_encoding(input_path)
    try:
        df = pd.read_csv(input_path, encoding=enc, sep=delimiter or ",")
    except (UnicodeDecodeError, pd.errors.ParserError):
        for fallback in ["utf-8", "latin-1", "iso-8859-1", "cp1252"]:
            try:
                df = pd.read_csv(input_path, encoding=fallback, sep=delimiter or ",")
                break
            except (UnicodeDecodeError, pd.errors.ParserError):
                continue
        else:
            raise ValueError("Could not decode CSV.")

    _dataframe_to_pdf(df, output_path, title="CSV Data")


def excel_to_csv(input_path: Path, output_path: Path) -> None:
    """Convert Excel to CSV. Supports .xlsx and .xls."""
    if input_path.suffix.lower() == ".xls":
        df = pd.read_excel(input_path, engine="xlrd")
    else:
        df = pd.read_excel(input_path, engine="openpyxl")
    df.to_csv(output_path, index=False, encoding="utf-8")


def pdf_to_excel(
    input_path: Path,
    output_path: Path,
    pages: list[int] | None = None,
    password: str | None = None,
    status_callback: callable = None,
) -> tuple[pd.DataFrame | None, str | None]:
    """
    Extract tables from PDF and save as Excel.
    """
    import pdfplumber

    all_tables = []
    warning = None
    kwargs = {"password": password} if password else {}

    with pdfplumber.open(input_path, **kwargs) as pdf:
        page_range = pages if pages else range(len(pdf.pages))
        total = len(page_range)
        for idx, i in enumerate(page_range):
            if i < 0 or i >= len(pdf.pages): continue
            if status_callback:
                status_callback(f"Analyzing Page {idx + 1} of {total}...", (idx / total))
            
            page = pdf.pages[i]
            tables = page.extract_tables()
            for tbl in tables:
                if tbl and any(any(cell for cell in row) for row in tbl):
                    cleaned = [[str(c).strip() if c else "" for c in row] for row in tbl]
                    all_tables.append(cleaned)

        if not all_tables:
            all_text_lines = []
            for i, page in enumerate(pdf.pages):
                if status_callback:
                    status_callback(f"Scanning Text - Page {i+1}...", (i / len(pdf.pages)))
                text = page.extract_text()
                if text and text.strip():
                    lines = [l.strip().split() for l in text.split("\n") if l.strip()]
                    all_text_lines.extend(lines)

            if all_text_lines:
                all_tables.append(all_text_lines)
                warning = "No structured tables found. Text was extracted as rows instead."
            else:
                raise ValueError("No extractable data found. This PDF might be scanned.")

    dfs = []
    for tbl in all_tables:
        if len(tbl) > 1:
            cols = _deduplicate_columns(tbl[0])
            df = pd.DataFrame(tbl[1:], columns=cols)
        elif tbl:
            df = pd.DataFrame(tbl)
            df.columns = _deduplicate_columns(df.columns)
        else:
            continue
        if not df.empty:
            dfs.append(df)

    if not dfs:
        raise ValueError("No valid table data could be extracted from the PDF.")

    combined = pd.concat(dfs, ignore_index=True)
    combined.to_excel(output_path, index=False, engine="openpyxl")
    _autofit_excel_columns(output_path)
    return combined, warning


def pdf_to_xml(
    input_path: Path,
    output_path: Path,
    pages: list[int] | None = None,
    password: str | None = None,
    status_callback: callable = None,
) -> tuple[str | None, str | None]:
    """
    High-precision extraction from PDFs (specialized for Bank Statements) to XML.
    """
    import pdfplumber
    import xml.etree.ElementTree as ET
    from xml.dom import minidom
    import re
    from datetime import datetime

    pages_data = []
    kwargs = {"password": password} if password else {}

    with pdfplumber.open(input_path, **kwargs) as pdf:
        page_range = pages if pages else range(len(pdf.pages))
        total = len(page_range)
        for idx, i in enumerate(page_range):
            if i < 0 or i >= len(pdf.pages): continue
            if status_callback:
                status_callback(f"Extracting XML Data - Page {idx + 1} of {total}...", (idx / total))
            page = pdf.pages[i]
            pages_data.append({
                "page": i + 1,
                "text": page.extract_text() or "",
                "tables": page.extract_tables() or []
            })

    root = ET.Element("BankStatement")
    root.set("source", input_path.name)
    root.set("extractedAt", datetime.now().isoformat())

    first_text = pages_data[0]["text"] if pages_data else ""
    account_el = ET.SubElement(root, "AccountInfo")
    
    # Standard patterns from the skill
    patterns = [
        ("AccountNumber", r'(?:account\s*(?:number|no\.?|#))\s*[:\-]?\s*(\d[\d\s\-*X]+)'),
        ("StatementPeriod", r'(?:statement\s*period|period)\s*[:\-]?\s*([^\n]{5,40})'),
        ("AccountHolder",   r'(?:name|account\s*holder)\s*[:\-]?\s*([A-Z][^\n]{2,50})'),
        ("IBAN",            r'\b([A-Z]{2}\d{2}[A-Z0-9]{4,30})\b'),
        ("Currency",        r'\b(USD|EUR|GBP|SAR|AED|EGP|KWD|BHD|QAR|OMR|JOD|TND|MAD|KES)\b'),
    ]
    
    for label, pat in patterns:
        m = re.search(pat, first_text, re.IGNORECASE)
        if m: ET.SubElement(account_el, label).text = m.group(1).strip()

    summary_el = ET.SubElement(root, "Summary")
    summary_pats = [
        ("OpeningBalance", r'(?:opening|beginning|prev(?:ious)?)\s*balance\s*[:\-]?\s*([\d,\.]+)'),
        ("ClosingBalance", r'(?:closing|ending|current)\s*balance\s*[:\-]?\s*([\d,\.]+)'),
        ("TotalCredits",   r'total\s*credits?\s*[:\-]?\s*([\d,\.]+)'),
        ("TotalDebits",    r'total\s*debits?\s*[:\-]?\s*([\d,\.]+)'),
    ]
    for label, pat in summary_pats:
        for page in pages_data:
            m = re.search(pat, page["text"], re.IGNORECASE)
            if m:
                ET.SubElement(summary_el, label).text = m.group(1).strip()
                break

    transactions_el = ET.SubElement(root, "Transactions")
    txn_count = 0
    for page in pages_data:
        for table in page["tables"]:
            if not table or len(table) < 2: continue
            header = [str(h).lower().strip() for h in table[0]]
            col_map = {}
            for i, h in enumerate(header):
                if any(k in h for k in ["date", "dt", "value date"]): col_map["date"] = i
                elif any(k in h for k in ["description", "narration", "detail", "particulars", "reference"]): col_map["description"] = i
                elif any(k in h for k in ["debit", "withdrawal", "dr"]): col_map["debit"] = i
                elif any(k in h for k in ["credit", "deposit", "cr"]): col_map["credit"] = i
                elif any(k in h for k in ["balance"]): col_map["balance"] = i
                elif "amount" in h and "debit" not in col_map and "credit" not in col_map: col_map["amount"] = i

            for row in table[1:]:
                if not any(row): continue
                txn_count += 1
                txn_el = ET.SubElement(transactions_el, "Transaction", id=str(txn_count))
                for field, idx in col_map.items():
                    if idx < len(row) and row[idx]:
                        ET.SubElement(txn_el, field.capitalize()).text = str(row[idx]).strip()

    xml_str = minidom.parseString(ET.tostring(root, encoding='unicode')).toprettyxml(indent="  ")
    output_path.write_text(xml_str, encoding="utf-8")
    return xml_str, None


def pdf_to_json(
    input_path: Path,
    output_path: Path,
    pages: list[int] | None = None,
    password: str | None = None,
    status_callback: callable = None,
) -> tuple[dict | None, str | None]:
    """Convert PDF data (text and tables) to structured JSON."""
    import pdfplumber
    import json

    data = {"source": input_path.name, "pages": []}
    kwargs = {"password": password} if password else {}

    with pdfplumber.open(input_path, **kwargs) as pdf:
        page_range = pages if pages else range(len(pdf.pages))
        total = len(page_range)
        for idx, i in enumerate(page_range):
            if i < 0 or i >= len(pdf.pages): continue
            if status_callback:
                status_callback(f"Parsing JSON Structure - Page {idx + 1} of {total}...", (idx / total))
            page = pdf.pages[i]
            data["pages"].append({
                "page": i + 1,
                "text": page.extract_text() or "",
                "tables": page.extract_tables() or []
            })

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    return data, None


def excel_to_pdf(input_path: Path, output_path: Path) -> None:
    """
    Convert Excel to PDF using pandas + fpdf2. No LibreOffice required.
    Supports .xlsx and .xls. All sheets are rendered as tables.
    """
    engine = "xlrd" if input_path.suffix.lower() == ".xls" else "openpyxl"
    xl = pd.ExcelFile(input_path, engine=engine)

    from fpdf import FPDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    sheet_count = 0
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet_name)
        if df.empty or len(df.columns) == 0:
            continue
        sheet_count += 1
        _add_dataframe_page(pdf, df, title=sheet_name)

    if sheet_count == 0:
        raise ValueError("Excel file has no data in any sheet.")
    pdf.output(str(output_path))


# --- JSON converters ---

def json_to_csv(input_path: Path, output_path: Path) -> None:
    """Convert JSON to CSV. Handles arrays and nested objects."""
    import json

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    df = _json_to_dataframe(data)
    df.to_csv(output_path, index=False, encoding="utf-8")


def json_to_excel(input_path: Path, output_path: Path) -> None:
    """Convert JSON to Excel. Handles arrays and nested objects."""
    import json

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    df = _json_to_dataframe(data)
    df.to_excel(output_path, index=False, engine="openpyxl")
    _autofit_excel_columns(output_path)


def json_to_pdf(input_path: Path, output_path: Path) -> None:
    """Convert JSON data to a formatted PDF table."""
    import json

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    df = _json_to_dataframe(data)
    _dataframe_to_pdf(df, output_path, title="JSON Data")


# --- XML converters ---

def xml_to_csv(input_path: Path, output_path: Path) -> None:
    """Convert XML to CSV using xmltodict for flattening."""
    import xmltodict

    with open(input_path, "r", encoding="utf-8") as f:
        data = xmltodict.parse(f.read())

    df = _nested_dict_to_dataframe(data)
    df.to_csv(output_path, index=False, encoding="utf-8")


def xml_to_excel(input_path: Path, output_path: Path) -> None:
    """Convert XML to Excel using xmltodict for flattening."""
    import xmltodict

    with open(input_path, "r", encoding="utf-8") as f:
        data = xmltodict.parse(f.read())

    df = _nested_dict_to_dataframe(data)
    df.to_excel(output_path, index=False, engine="openpyxl")
    _autofit_excel_columns(output_path)


def xml_to_json(input_path: Path, output_path: Path) -> None:
    """Convert XML to JSON."""
    import json
    import xmltodict

    with open(input_path, "r", encoding="utf-8") as f:
        data = xmltodict.parse(f.read())

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# --- YAML converters ---

def yaml_to_json(input_path: Path, output_path: Path) -> None:
    """Convert YAML to JSON."""
    import json
    import yaml

    with open(input_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def yaml_to_csv(input_path: Path, output_path: Path) -> None:
    """Convert YAML to CSV. Best for YAML arrays of objects."""
    import yaml

    with open(input_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    df = _json_to_dataframe(data)
    df.to_csv(output_path, index=False, encoding="utf-8")


def yaml_to_excel(input_path: Path, output_path: Path) -> None:
    """Convert YAML to Excel. Best for YAML arrays of objects."""
    import yaml

    with open(input_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    df = _json_to_dataframe(data)
    df.to_excel(output_path, index=False, engine="openpyxl")
    _autofit_excel_columns(output_path)


def json_to_yaml(input_path: Path, output_path: Path) -> None:
    """Convert JSON to YAML."""
    import json
    import yaml

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    with open(output_path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)


# --- Helpers ---

def _autofit_excel_columns(output_path: Path) -> None:
    """Adjust column widths in Excel for readability."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(output_path)
        ws = wb.active
        for i, col in enumerate(ws.columns, 1):
            max_len = min(max(len(str(c.value) or "") for c in col) + 2, 50)
            ws.column_dimensions[get_column_letter(i)].width = max_len
        wb.save(output_path)
    except Exception:
        pass  # Non-critical


def _to_latin1(s: str, max_len: int = 80) -> str:
    """Encode to latin-1 for fpdf2 compatibility."""
    text = str(s)[:max_len]
    return text.encode("latin-1", "replace").decode("latin-1")


def _json_to_dataframe(data) -> pd.DataFrame:
    """Convert JSON data (list or dict) to a DataFrame."""
    if isinstance(data, list):
        df = pd.json_normalize(data, max_level=2)
    elif isinstance(data, dict):
        # Try to find the main array in the dict
        for key, val in data.items():
            if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                df = pd.json_normalize(val, max_level=2)
                break
        else:
            # Flat dict - single row
            df = pd.json_normalize(data, max_level=2)
    else:
        raise ValueError("JSON must be an object or array.")

    if df.empty:
        raise ValueError("No tabular data could be extracted from JSON.")
    return df


def _nested_dict_to_dataframe(data: dict) -> pd.DataFrame:
    """Flatten nested XML dict to DataFrame."""
    # xmltodict wraps in root element; drill down to find arrays
    def _find_records(d, depth=0):
        if depth > 5:
            return None
        if isinstance(d, list):
            return d
        if isinstance(d, dict):
            for k, v in d.items():
                result = _find_records(v, depth + 1)
                if result is not None:
                    return result
        return None

    records = _find_records(data)
    if records and isinstance(records, list):
        df = pd.json_normalize(records, max_level=2)
    elif isinstance(data, dict):
        df = pd.json_normalize(data, max_level=3)
    else:
        raise ValueError("Could not extract tabular data from XML.")

    if df.empty:
        raise ValueError("No data could be extracted from XML.")
    return df


def _dataframe_to_pdf(df: pd.DataFrame, output_path: Path, title: str = "Data") -> None:
    """Render a DataFrame as a PDF table."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    _add_dataframe_page(pdf, df, title=title)
    pdf.output(str(output_path))


def _add_dataframe_page(pdf, df: pd.DataFrame, title: str = "") -> None:
    """Add a DataFrame as a table page to an existing FPDF object."""
    col_count = len(df.columns)
    font_size = max(6, 10 - col_count // 5)
    orientation = "L" if col_count > 8 else "P"
    pdf.set_font("Helvetica", size=font_size)
    pdf.add_page(orientation=orientation)

    if title:
        pdf.set_font("Helvetica", "B", font_size + 2)
        pdf.cell(0, 8, _to_latin1(title, 100), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=font_size)
        pdf.ln(2)

    headers = [_to_latin1(str(c)) for c in df.columns]
    rows = [
        [_to_latin1(str(v) if pd.notna(v) else "") for v in row]
        for row in df.values.tolist()
    ]
    table_data = [headers] + rows
    with pdf.table(
        table_data,
        text_align="L",
        cell_fill_color=235,
        cell_fill_mode="ROWS",
    ):
        pass
